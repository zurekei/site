#!/usr/bin/env python3
"""data/births_forecast.csv と data/births_actual.csv を一次資料から作り直す。

出生数(人数)の「歴代推計 vs 実績」。合計特殊出生率(比率)を扱う
data/fertility_*.csv とは別ファイルで、指標としても別ページ(/births)にする。

■ 推計側 — 国立社会保障・人口問題研究所「日本の将来推計人口」の各版
  どの版も「出生，死亡および自然増加の実数ならびに率」という同じ表を持つ。
  版によって表番号もファイル形式も違う(下の SOURCES)ので、URLは版ごとに
  手で押さえたうえで、**数値はスクリプトが原本から読む**。転記しない。

  ⚠ 基準が版で2種類ある。ここが唯一の落とし穴で、無視すると1〜2%の系統的な
  ズレを予測誤差と取り違える。
    - 1992〜2012年推計: 表の脚注に「日本における外国人を含む．」とある
      (1992年推計だけは脚注が無いが、報告書の「2.基準人口」が国勢調査ベースの
      総務庁統計局推計人口=総人口と書いている)。→ forecast_basis=total
    - 2017・2023年推計: 総人口ベースの表1-8とは別に、日本人人口ベースの
      表1-8(J)がある。実績(人口動態統計=日本における日本人)と基準が揃うので
      **(J)の方を採る**。→ forecast_basis=japanese
  実績側は日本人ベースしか年次の完全な系列が無いため(§後述)、2012年以前の
  4版だけは基準が食い違ったまま並ぶ。隠さず forecast_basis 列に出して、
  ページ側でも版ごとに注記する。差の実測値は2021年で
  総人口831.001千人 - 日本人813.120千人 = 17.9千人(2.2%)。

■ 実績側 — 厚生労働省「人口動態統計」第2表 人口動態総覧の年次推移
  1980〜2024年は令和6年(2024)確定数の統計表、2025年は令和7年(2025)
  月報年計(概数)の統計表から読む。どちらも同じ第2表(第1表)で、
  「日本における日本人」について。

■ 1992年推計だけ原本がスキャンPDFで、機械的に読めない
  報告書(101827_4.pdf)48ページの表2-5を目視で書き写している。ここだけは
  転記なので、出生・死亡・自然増加の3列すべてを持たせて
  「出生 - 死亡 = 自然増加」が全35行で成り立つことをこのスクリプトが検算する
  (3値とも1000人単位に丸めた値なので、独立した丸めの分だけ±1のずれは出る。
  2以上ずれたら桁の読み違いとみなして止める)。数字を1つ読み違えれば
  まず通らないので、目視の転記としてはこれで担保する。

使い方:
    python3 bin/build_births.py            # 原本を取得して2つのCSVを作り直す
    python3 bin/build_births.py --dry-run  # 標準出力に出すだけ

原本はネットワークから取り直す。更新は年1回の手動運用が前提で(../CLAUDE.md)、
ビルドの経路には入っていない。
"""
import argparse
import csv
import io
import os
import re
import sys
import urllib.request
from decimal import Decimal

UA = "zurekei-build/1.0 (+https://zurekei.org/)"
SITE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
IPSS = "https://www.ipss.go.jp"
MHLW = "https://www.mhlw.go.jp"

# ── 推計側 ────────────────────────────────────────────────────────────────
# url        … その版の出生数表そのもの(この URL が CSV の出典列に入る)
# published  … その版の報告書の公表日。data/fertility_forecast.csv の
#              forecast_published_date と同じ値を使う(同じ推計なので揃える)
# basis      … total=日本における外国人を含む / japanese=日本における日本人
# years      … 本推計の対象期間。参考推計(超長期)の行は採らない
# note       … CSV の notes 列(カンマ不可。csv.js のパーサが単純split)
SOURCES = {
    1997: {
        "url": f"{IPSS}/pp-zenkoku/j/zenkoku1997/db_zenkoku1997/s_tables/pp1997_chu_T2_1_6.xls",
        "published": "1997-04-25",
        "basis": "total",
        "years": (1996, 2050),
        "note": "表2-1-6(中位推計)。1000人単位の整数で公表されており1000人未満は丸められている",
    },
    2002: {
        "url": f"{IPSS}/pp-zenkoku/j/zenkoku2002/db_zenkoku2002/s_tables/pp2002_chu_T1_1_7.xls",
        "published": "2002-01",
        "basis": "total",
        "years": (2001, 2050),
        "note": "表1-1-7(中位推計)",
    },
    2006: {
        "url": f"{IPSS}/syoushika/tohkei/suikei07/houkoku/kekka-1/1-8.xls",
        "published": "2006-12",
        "basis": "total",
        "years": (2006, 2055),
        "note": "表1-8(出生中位・死亡中位推計)",
    },
    2012: {
        "url": f"{IPSS}/syoushika/tohkei/newest04/s-kekka/1-8.xls",
        "published": "2012-01",
        "basis": "total",
        "years": (2011, 2060),
        "note": "表1-8(出生中位・死亡中位推計)",
    },
    2017: {
        "url": f"{IPSS}/pp-zenkoku/j/zenkoku2017/db_zenkoku2017/s_tables/1-8(J).xls",
        "published": "2017-04",
        "basis": "japanese",
        "years": (2016, 2065),
        "note": "表1-8(J)(日本人人口・出生中位・死亡中位推計)",
    },
    2023: {
        "url": f"{IPSS}/pp-zenkoku/j/zenkoku2023/db_zenkoku2023/s_tables/1-8(J).xlsx",
        "published": "2023-04-26",
        "basis": "japanese",
        "years": (2021, 2070),
        "note": "表1-8(J)(日本人人口・出生中位・死亡中位推計)",
    },
}

# 1992年推計。原本がスキャンPDFのため、ここだけ目視で書き写している。
# 報告書48ページ「表2-5 出生，死亡および自然増加の実数ならびに率:中位推計」の
# 実数欄(単位1000人)。(年, 出生, 死亡, 自然増加)。下の check_1992() が
# 出生-死亡=自然増加 を全行で検算する。
V1992_URL = f"{IPSS}/syoushika/bunken/data/pdf/101827_4.pdf"
V1992_PUBLISHED = "1992-09-30"
V1992_NOTE = (
    "表2-5(中位推計)。原本はスキャンPDFのため数値は目視で書き写しており"
    "1000人単位の整数で公表されている。出生-死亡=自然増加が全35年分で"
    "成り立つことをbin/build_births.pyが検算している"
)
V1992_ROWS = [
    (1991, 1231, 834, 397),
    (1992, 1233, 866, 367),
    (1993, 1236, 882, 353),
    (1994, 1248, 900, 349),
    (1995, 1269, 918, 352),
    (1996, 1298, 937, 361),
    (1997, 1332, 958, 373),
    (1998, 1368, 980, 388),
    (1999, 1404, 1002, 402),
    (2000, 1438, 1026, 413),
    (2001, 1467, 1050, 417),
    (2002, 1489, 1077, 411),
    (2003, 1503, 1106, 397),
    (2004, 1508, 1135, 373),
    (2005, 1506, 1166, 340),
    (2006, 1497, 1197, 300),
    (2007, 1482, 1229, 253),
    (2008, 1463, 1261, 202),
    (2009, 1440, 1293, 147),
    (2010, 1415, 1327, 88),
    (2011, 1388, 1358, 30),
    (2012, 1360, 1389, -29),
    (2013, 1331, 1419, -87),
    (2014, 1303, 1447, -145),
    (2015, 1274, 1476, -201),
    (2016, 1247, 1502, -255),
    (2017, 1222, 1527, -305),
    (2018, 1199, 1550, -352),
    (2019, 1179, 1572, -394),
    (2020, 1162, 1594, -432),
    (2021, 1150, 1615, -465),
    (2022, 1143, 1636, -493),
    (2023, 1139, 1655, -516),
    (2024, 1140, 1674, -534),
    (2025, 1144, 1694, -550),
]

# ── 実績側 ────────────────────────────────────────────────────────────────
ACTUAL_CONFIRMED = {
    "url": f"{MHLW}/toukei/saikin/hw/jinkou/kakutei24/xls/16_hyoR06.xlsx",
    # 第2表は年代でシートが4枚に割れていて、1960年以降は③(①は1899〜1959年)
    "sheet": "第２表③",
    "years": (1980, 2024),
    "note": "",
}
ACTUAL_PRELIM = {
    "url": f"{MHLW}/toukei/saikin/hw/jinkou/geppo/nengai25/xls/R7toukeihyou.xlsx",
    "sheet": "第1表",
    "years": (2025, 2025),
    "note": (
        "[実績]厚生労働省「令和7(2025)年 人口動態統計月報年計(概数)の概況」統計表 第1表より。"
        "確定数ではなく概数の段階の値で、例年9月公表の確定数で改定される見込み"
    ),
}
ACTUAL_HEAD_NOTE = "[実績]厚生労働省「令和6(2024)年 人口動態統計(確定数)の概況」統計表 第2表より。日本における日本人について"


def fetch(url):
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=120) as r:
        return r.read()


def rows_of(blob, sheet=None):
    """xls / xlsx のどちらでも、シートを行ごとの文字列リストにして返す。"""
    if blob[:2] == b"PK":
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        return [
            ["" if c is None else str(c) for c in row]
            for row in ws.iter_rows(values_only=True)
        ]
    import xlrd

    wb = xlrd.open_workbook(file_contents=blob)
    ws = wb.sheet_by_name(sheet) if sheet else wb.sheet_by_index(0)
    return [
        [str(ws.cell_value(r, c)) for c in range(ws.ncols)] for r in range(ws.nrows)
    ]


YEAR_CELL = re.compile(r"^\(?((?:19|20|21)\d\d)\)?(?:\.0)?$")
NUM_CELL = re.compile(r"^-?\d+(?:\.\d+)?$")


def read_projection(blob):
    """表1-8系の「年次 → 出生数(1000人単位の文字列)」。列の並びは版で違うが、
    どの版も『年次の右にある最初の数値が出生』という並びは同じ。"""
    out = {}
    for row in rows_of(blob):
        year = None
        yi = None
        for i, cell in enumerate(row[:3]):
            m = YEAR_CELL.match(cell.strip())
            if m:
                year, yi = int(m.group(1)), i
                break
        if year is None:
            continue
        for cell in row[yi + 1 :]:
            cell = cell.strip().replace(",", "")
            if NUM_CELL.match(cell):
                out[year] = cell
                break
    return out


def thousands_to_persons(text):
    """「1059.245」(1000人単位) → 1059245(人)。floatを経由しない。"""
    return int(Decimal(text) * 1000)


BIRTHS_HEAD = re.compile(r"出\s*生\s*数")


def read_actual(blob, sheet):
    """人口動態総覧の年次推移から「年次 → 出生数(人)」。

    ⚠ 「年次の右にある最初の数値」では読めない。年次欄の隣は和暦の年で、
    それも数値だから拾ってしまう(実際に最初こう書いて全行が和暦を出生数として
    読んだ)。列は見出しから決める:
      - 出生数の列 … 見出しに「出生数」とある列
      - 年次の列   … その左側で、西暦4桁が最も多く並ぶ列
    年次欄は10年ごとにだけ4桁で、あいだの行は下1〜2桁だけ、という体裁の版が
    あるので(確定数の統計表がこれ)、直前に読んだ4桁から復元する。
    """
    grid = [[c.strip() for c in row] for row in rows_of(blob, sheet)]
    bi = None
    for row in grid[:12]:
        for i, cell in enumerate(row):
            if BIRTHS_HEAD.search(cell):
                bi = i
                break
        if bi is not None:
            break
    if bi is None:
        raise SystemExit(f"実績: シート{sheet}に「出生数」の見出しが見つからない")

    counts = [
        sum(1 for row in grid if i < len(row) and YEAR_CELL.match(row[i]))
        for i in range(bi)
    ]
    if not counts or max(counts) == 0:
        raise SystemExit(f"実績: シート{sheet}に年次の列が見つからない")
    yi = counts.index(max(counts))

    out = {}
    last4 = None
    for row in grid:
        if yi >= len(row) or bi >= len(row):
            continue
        cell = row[yi]
        m = YEAR_CELL.match(cell)
        if m:
            year = int(m.group(1))
        elif last4 is not None and re.fullmatch(r"\d{1,2}", cell):
            year = last4 - (last4 % 100) + int(cell)
            if year <= last4:
                year += 100
            if year - last4 > 10:
                continue
        else:
            continue
        births = row[bi].replace(",", "")
        if not NUM_CELL.match(births):
            continue
        last4 = year
        out[year] = int(Decimal(births))
    return out


def check_1992():
    bad = []
    for year, birth, death, natural in V1992_ROWS:
        if abs((birth - death) - natural) > 1:
            bad.append(f"{year}: {birth} - {death} != {natural}")
    if bad:
        raise SystemExit(
            "1992年推計の転記が「出生-死亡=自然増加」を満たしていない:\n  "
            + "\n  ".join(bad)
        )
    years = [r[0] for r in V1992_ROWS]
    if years != list(range(1991, 2026)):
        raise SystemExit(f"1992年推計の年次が1991〜2025年で連続していない: {years}")


def build_forecast():
    check_1992()
    rows = []
    for year, birth, _death, _natural in V1992_ROWS:
        rows.append(
            {
                "vintage_year": 1992,
                "target_year": year,
                "projected_births": birth * 1000,
                "forecast_basis": "total",
                "forecast_source_url": V1992_URL,
                "forecast_published_date": V1992_PUBLISHED,
                "notes": V1992_NOTE,
            }
        )
    for vintage in sorted(SOURCES):
        cfg = SOURCES[vintage]
        data = read_projection(fetch(cfg["url"]))
        lo, hi = cfg["years"]
        got = [y for y in sorted(data) if lo <= y <= hi]
        if got != list(range(lo, hi + 1)):
            raise SystemExit(
                f"{vintage}年推計: {lo}〜{hi}年が揃っていない(読めたのは{len(got)}年分)"
            )
        for y in got:
            rows.append(
                {
                    "vintage_year": vintage,
                    "target_year": y,
                    "projected_births": thousands_to_persons(data[y]),
                    "forecast_basis": cfg["basis"],
                    "forecast_source_url": cfg["url"],
                    "forecast_published_date": cfg["published"],
                    "notes": cfg["note"],
                }
            )
        print(
            f"  {vintage}年推計 {len(got)}年分 ({lo}〜{hi}) basis={cfg['basis']}",
            file=sys.stderr,
        )
    return rows


def build_actual():
    rows = []
    for cfg, head_note in ((ACTUAL_CONFIRMED, ACTUAL_HEAD_NOTE), (ACTUAL_PRELIM, None)):
        data = read_actual(fetch(cfg["url"]), cfg["sheet"])
        lo, hi = cfg["years"]
        got = [y for y in sorted(data) if lo <= y <= hi]
        if got != list(range(lo, hi + 1)):
            missing = sorted(set(range(lo, hi + 1)) - set(got))
            raise SystemExit(f"実績: {lo}〜{hi}年のうち{missing}が読めなかった")
        for y in got:
            note = cfg["note"]
            if head_note and y == lo:
                note = head_note
            rows.append(
                {
                    "year": y,
                    "actual_births": data[y],
                    "source_url": cfg["url"],
                    "notes": note,
                }
            )
        print(f"  実績 {len(got)}年分 ({lo}〜{hi})", file=sys.stderr)
    return rows


def write(path, cols, rows, dry_run):
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=cols, lineterminator="\n")
    w.writeheader()
    for r in rows:
        for k, v in r.items():
            if isinstance(v, str) and "," in v:
                raise SystemExit(f"{path}: {k} にカンマが入っている(csv.jsが読めない): {v}")
        w.writerow(r)
    text = buf.getvalue()
    if dry_run:
        sys.stdout.write(text)
    else:
        with open(os.path.join(SITE_DIR, path), "w", encoding="utf-8") as f:
            f.write(text)
    print(f"  → {path} ({len(rows)}行)", file=sys.stderr)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    print("推計側:", file=sys.stderr)
    forecast = build_forecast()
    print("実績側:", file=sys.stderr)
    actual = build_actual()

    write(
        "data/births_forecast.csv",
        [
            "vintage_year",
            "target_year",
            "projected_births",
            "forecast_basis",
            "forecast_source_url",
            "forecast_published_date",
            "notes",
        ],
        forecast,
        args.dry_run,
    )
    write(
        "data/births_actual.csv",
        ["year", "actual_births", "source_url", "notes"],
        actual,
        args.dry_run,
    )

    vintages = sorted({r["vintage_year"] for r in forecast})
    jp = sorted({r["vintage_year"] for r in forecast if r["forecast_basis"] == "japanese"})
    print(
        f"完了: 推計{len(vintages)}版({vintages}) うち日本人ベース{jp} / "
        f"実績{actual[0]['year']}〜{actual[-1]['year']}年",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
